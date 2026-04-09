# -*- coding: utf-8 -*-
"""
Point HVAC - Excel → JSON Donusturucu
Tum Excel dosyalarini www/data/ altina JSON olarak yazar.
Kullanim: python tools/build_data.py
"""

import json
import os
import sys
from openpyxl import load_workbook

# Yollar
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(PROJECT_DIR, 'www', 'data')

# Orijinal Excel dosyalarinin yolu
EXCEL_DIR = os.path.join(
    os.path.dirname(PROJECT_DIR),
    'point_hvac_android', 'point_hvac', 'files'
)


def to_float(x):
    """Guvenli sayi donusumu. Turkce formati destekler."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(',', '.')
    cleaned = ''.join(ch for ch in s if (ch.isdigit() or ch in '.-'))
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def to_str(x):
    """Guvenli string donusumu."""
    if x is None:
        return ''
    return str(x).strip()


def convert_fan_modeller(excel_path, out_dir):
    """fan_modeller.xlsx → Her sheet icin ayri JSON."""
    print(f'  fan_modeller.xlsx ...')
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    os.makedirs(out_dir, exist_ok=True)

    index = {'sheets': []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Ilk satir: basinc degerleri (B1, C1, D1, ...)
        header = rows[0]
        basinc_values = []
        for v in header[1:]:
            bv = to_float(v)
            if bv is not None:
                basinc_values.append(bv)
            else:
                break  # Bos hucreye ulasinca dur

        if not basinc_values:
            continue

        # Veri satirlari: model kodu + debi degerleri
        models = []
        for row in rows[1:]:
            model = to_str(row[0])
            if not model:
                continue
            debi_values = []
            for i, bv in enumerate(basinc_values):
                dv = to_float(row[i + 1]) if (i + 1) < len(row) else None
                debi_values.append(dv)
            models.append({
                'model': model,
                'debi_values': debi_values
            })

        sheet_data = {
            'sheet_name': sheet_name,
            'basinc_values': basinc_values,
            'models': models
        }

        safe_name = sheet_name.replace(' ', '_').replace('/', '_').lower()
        out_path = os.path.join(out_dir, f'{safe_name}.json')
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(sheet_data, f, ensure_ascii=False, indent=1)

        index['sheets'].append({
            'name': sheet_name,
            'file': f'{safe_name}.json',
            'model_count': len(models),
            'basinc_count': len(basinc_values)
        })
        print(f'    {sheet_name}: {len(models)} model, {len(basinc_values)} basinc')

    with open(os.path.join(out_dir, '_index.json'), 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=1)

    wb.close()


def convert_fan_fiyatlar(excel_path, out_path):
    """fan_fiyatlar.xlsx → JSON."""
    print(f'  fan_fiyatlar.xlsx ...')
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        items = []
        for row in rows[1:]:  # Header atla
            if not row or not row[0]:
                continue
            item = {
                'model': to_str(row[0]),
                'fiyat': to_float(row[1]) if len(row) > 1 else None,
                'kontrol_ekipman': to_str(row[2]) if len(row) > 2 else '',
                'motor_gucu': to_float(row[3]) if len(row) > 3 else None,
                'voltaj': to_str(row[4]) if len(row) > 4 else '',
                'ses_seviyesi': to_float(row[5]) if len(row) > 5 else None,
                'devir': to_float(row[6]) if len(row) > 6 else None,
            }
            items.append(item)
        result[sheet_name] = items
        print(f'    {sheet_name}: {len(items)} model')

    wb.close()
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=1)


def convert_simple_excel(excel_path, out_path, columns, label=''):
    """Basit Excel → JSON (model, debi, fiyat gibi sabit kolonlu)."""
    print(f'  {label or os.path.basename(excel_path)} ...')
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        items = []
        for row in rows[1:]:  # Header atla
            if not row or not row[0]:
                continue
            item = {}
            for field, idx, conv in columns:
                if idx < len(row):
                    item[field] = conv(row[idx])
                else:
                    item[field] = None
            items.append(item)
        result[sheet_name] = items
        print(f'    {sheet_name}: {len(items)} kayit')

    wb.close()
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=1)


def convert_aksa(excel_path, out_path):
    """AKSA_MOTOR_FAN_2026_FIYAT.xlsx → JSON (farkli kolon yapilari)."""
    print(f'  AKSA fiyat ...')
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    result = {}

    # Sheet isimleri ve kolon haritalari
    COLUMN_MAPS = {
        'default': [
            ('kategori', 0, to_str), ('model', 1, to_str), ('seri', 2, to_str),
            ('tip', 3, to_str), ('voltaj', 4, to_str), ('guc_w', 5, to_float),
            ('debi', 6, to_float), ('devir', 7, to_float), ('ses_db', 8, to_float),
            ('koli', 9, to_str), ('fiyat', 10, to_float)
        ],
        'sogutma': [
            ('kategori', 0, to_str), ('model', 1, to_str), ('seri', 2, to_str),
            ('voltaj', 3, to_str), ('akim_a', 4, to_float), ('guc_w', 5, to_float),
            ('debi', 6, to_float), ('devir', 7, to_float), ('ses_db', 8, to_float),
            ('fiyat', 9, to_float)
        ],
        'ec': [
            ('kategori', 0, to_str), ('model', 1, to_str), ('tip', 2, to_str),
            ('voltaj', 3, to_str), ('guc_w', 4, to_float), ('debi', 5, to_float),
            ('devir', 6, to_float), ('basinc_pa', 7, to_float), ('ses_db', 8, to_float),
            ('koli', 9, to_str), ('fiyat', 10, to_float)
        ],
        '688': [
            ('kategori', 0, to_str), ('model', 1, to_str), ('guc_w', 2, to_float),
            ('fiyat_ayaksiz', 3, to_float), ('fiyat_ayakli', 4, to_float)
        ]
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Kolon haritasini belirle (sheet ismine gore)
        sn_lower = sheet_name.lower()
        if '688' in sn_lower:
            cmap = COLUMN_MAPS['688']
        elif 'ec' in sn_lower:
            cmap = COLUMN_MAPS['ec']
        elif 'sogut' in sn_lower or 'soğut' in sn_lower or 'motor' in sn_lower.replace('fan', ''):
            cmap = COLUMN_MAPS['sogutma']
        else:
            cmap = COLUMN_MAPS['default']

        items = []
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            item = {}
            for field, idx, conv in cmap:
                if idx < len(row):
                    item[field] = conv(row[idx])
                else:
                    item[field] = None
            items.append(item)

        result[sheet_name] = {'column_map': cmap[0][0], 'items': items}
        print(f'    {sheet_name}: {len(items)} kayit')

    wb.close()
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=1)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    fan_dir = os.path.join(OUTPUT_DIR, 'fan_modeller')

    print('Point HVAC - Excel -> JSON Donusturucu')
    print(f'Kaynak: {EXCEL_DIR}')
    print(f'Hedef:  {OUTPUT_DIR}')
    print()

    if not os.path.exists(EXCEL_DIR):
        print(f'HATA: Excel dizini bulunamadi: {EXCEL_DIR}')
        sys.exit(1)

    # 1. Fan modelleri (en buyuk — her sheet ayri JSON)
    convert_fan_modeller(
        os.path.join(EXCEL_DIR, 'fan_modeller.xlsx'),
        fan_dir
    )

    # 2. Fan fiyatlari
    convert_fan_fiyatlar(
        os.path.join(EXCEL_DIR, 'fan_fiyatlar.xlsx'),
        os.path.join(OUTPUT_DIR, 'fan_fiyatlar.json')
    )

    # 3. Santral modelleri
    convert_simple_excel(
        os.path.join(EXCEL_DIR, 'santral_modeller.xlsx'),
        os.path.join(OUTPUT_DIR, 'santral_modeller.json'),
        [('model', 0, to_str), ('debi', 1, to_float), ('fiyat', 2, to_float)],
        'santral_modeller.xlsx'
    )

    # 4. Motor gucleri
    convert_simple_excel(
        os.path.join(EXCEL_DIR, 'motor_gucleri.xlsx'),
        os.path.join(OUTPUT_DIR, 'motor_gucleri.json'),
        [('santral_adi', 1, to_str), ('islev', 2, to_str),
         ('guc', 3, to_float), ('adet', 4, to_float)],
        'motor_gucleri.xlsx'
    )

    # 5. Plug Fan (Hucreli)
    convert_simple_excel(
        os.path.join(EXCEL_DIR, 'plug_fan.xlsx'),
        os.path.join(OUTPUT_DIR, 'plug_fan.json'),
        [('model', 0, to_str), ('debi', 1, to_float), ('basinc', 2, to_float),
         ('fiyat', 3, to_float), ('g2_fiyat', 4, to_float),
         ('g4_fiyat', 5, to_float), ('motor', 6, to_float)],
        'plug_fan.xlsx'
    )

    # 6. Isi Geri Kazanim (Cihaz modulu)
    convert_simple_excel(
        os.path.join(EXCEL_DIR, 'isi_geri_kazanim.xlsx'),
        os.path.join(OUTPUT_DIR, 'isi_geri_kazanim.json'),
        [('model', 0, to_str), ('debi', 1, to_float), ('fiyat', 2, to_float)],
        'isi_geri_kazanim.xlsx'
    )

    # 7. AKSA Motor & Fan
    convert_aksa(
        os.path.join(EXCEL_DIR, 'AKSA_MOTOR_FAN_2026_FIYAT.xlsx'),
        os.path.join(OUTPUT_DIR, 'aksa_fiyat.json')
    )

    print()
    print('Tamamlandi!')

    # Boyut raporu
    total = 0
    for root, dirs, files in os.walk(OUTPUT_DIR):
        for f in files:
            fp = os.path.join(root, f)
            sz = os.path.getsize(fp)
            total += sz
    print(f'Toplam JSON boyutu: {total / 1024:.1f} KB')


if __name__ == '__main__':
    main()
